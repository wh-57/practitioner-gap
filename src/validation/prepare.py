"""
Prepare the validation audit: pick a stratified sample of documents and
build the Excel audit workbook in one pass.

Reads:
    output/documents_deep.csv
    output/citations_deep.csv

Writes:
    output/validation/sample.csv            (used later by analyze.py)
    output/validation/audit_template.xlsx   (the workbook you fill in)

Then do the human audit, save the workbook, and run analyze.py.
"""

import argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Paths anchored to script location: gap/src/validation/prepare.py
REPO_ROOT = Path(__file__).resolve().parent.parent.parent
DOC_PATH    = REPO_ROOT / "output" / "documents_deep.csv"
CIT_PATH    = REPO_ROOT / "output" / "citations_deep.csv"
OUT_DIR     = REPO_ROOT / "output" / "validation"
SAMPLE_PATH = OUT_DIR / "sample.csv"
XLSX_PATH   = OUT_DIR / "audit_template.xlsx"

# Config
DEFAULT_N         = 50
MIN_PER_SOURCE    = 2
OVERSAMPLE_SCAN   = 5
OVERSAMPLE_TRUNC  = 5
FN_BLANKS_PER_DOC = 10
RANDOM_SEED       = 20260418


# ── Stratified sampling ──────────────────────────────────────────────────────

def _year_bin(y):
    if pd.isna(y):
        return "unknown"
    try:
        y = int(y)
    except (ValueError, TypeError):
        return "unknown"
    if y < 2000: return "pre_2000"
    if y < 2010: return "2000_2009"
    if y < 2020: return "2010_2019"
    return "2020_plus"


def build_sample(docs: pd.DataFrame, n: int, seed: int) -> pd.DataFrame:
    """Stratified sample: oversample scan/trunc, ensure min per source, fill remainder."""
    docs = docs.copy()
    year_col = "source_year_final" if "source_year_final" in docs.columns else "source_year"
    docs["year_bin"] = docs[year_col].apply(_year_bin)

    auditable = docs[docs["llm_status"] == "ok"].copy()
    if len(auditable) == 0:
        raise ValueError("No auditable documents (llm_status=ok) in documents_deep.csv")

    selected_ids = set()
    selected = []

    def add_from(subset, label, max_count, local_seed):
        candidates = subset[~subset["doc_id"].isin(selected_ids)]
        if len(candidates) == 0 or max_count <= 0:
            return
        picks = candidates.sample(min(len(candidates), max_count), random_state=local_seed)
        for _, row in picks.iterrows():
            selected_ids.add(row["doc_id"])
            r = dict(row)
            r["stratum"] = label
            selected.append(r)

    # 1. Oversample high-risk strata
    scanned = auditable[auditable["pdf_extraction_method"] == "likely_scanned"]
    add_from(scanned, "oversample_scanned", OVERSAMPLE_SCAN, seed + 1)

    truncated = auditable[auditable["doc_text_truncated"].astype(bool) == True]
    add_from(truncated, "oversample_truncated", OVERSAMPLE_TRUNC, seed + 2)

    # 2. Minimum per source type
    for src, grp in auditable.groupby("source_type"):
        add_from(grp, f"src_min_{src}", MIN_PER_SOURCE, seed + 3 + hash(src) % 10000)

    # 3. Fill remainder
    remaining = n - len(selected)
    if remaining > 0:
        pool = auditable[~auditable["doc_id"].isin(selected_ids)]
        if len(pool) > 0:
            picks = pool.sample(min(len(pool), remaining), random_state=seed + 4)
            for _, row in picks.iterrows():
                r = dict(row)
                r["stratum"] = f"fill_{row.get('source_type', 'unknown')}"
                selected.append(r)

    result = pd.DataFrame(selected)
    sort_cols = [c for c in ["source_type", year_col] if c in result.columns]
    if sort_cols:
        result = result.sort_values(sort_cols, na_position="last").reset_index(drop=True)
    result["doc_idx"] = range(1, len(result) + 1)
    return result


# ── Excel template ───────────────────────────────────────────────────────────

SUBFIELD_OPTIONS = [
    "asset_pricing", "corporate_finance", "financial_intermediation",
    "behavioral_finance", "market_microstructure", "macro_finance",
    "other_academic", "not_academic",
]
FUNCTION_OPTIONS = [
    "method_input", "empirical_evidence", "investment_rationale",
    "background_historical", "canonical_reference", "critique", "decorative",
]
VERDICT_OPTIONS = [
    "correct", "wrong_identification", "hallucinated", "human_found",
]

HEADERS = [
    # Context
    "doc_idx", "doc_id", "row_type", "citation_idx", "source_title", "source_file",
    # LLM extraction
    "raw_mention", "recovered_title", "recovered_authors", "recovered_year",
    "recovered_journal", "is_academic", "is_canonical", "academic_subfield",
    "citation_function", "citation_polarity", "confidence",
    # Human input
    "human_verdict",
    "human_correct_title", "human_correct_year", "human_correct_authors",
    "human_is_academic", "human_academic_subfield",
    "human_is_canonical", "human_citation_function",
    "notes",
]

HEADER_FILL    = PatternFill("solid", start_color="1F4E78")
HEADER_FONT    = Font(bold=True, color="FFFFFF")
LLM_FILL       = PatternFill("solid", start_color="E7EFF7")
FN_FILL        = PatternFill("solid", start_color="FFF2CC")
INPUT_FILL     = PatternFill("solid", start_color="FFFFFF")
SEPARATOR_FILL = PatternFill("solid", start_color="333333")

_INPUT_START = HEADERS.index("human_verdict") + 1
INPUT_COLUMNS = list(range(_INPUT_START, len(HEADERS) + 1))


def _safe(x):
    if x is None:
        return ""
    try:
        if pd.isna(x): return ""
    except (TypeError, ValueError):
        pass
    return x


def _build_instructions(ws):
    lines = [
        ("Validation Audit Template", Font(bold=True, size=18)),
        ("", None),
        ("How to use", Font(bold=True, size=14)),
        ("1. Go to the 'audit' sheet. Filter / sort by doc_idx to work one document at a time.", None),
        ("", None),
        ("2. Open the source PDF at the path in the source_file column.", None),
        ("", None),
        ("3. Walk through each LLM-extracted citation (blue rows):", None),
        ("     correct              — citation exists and is correctly identified", None),
        ("     wrong_identification — citation exists but LLM misidentified it", None),
        ("                            (fill human_correct_title / year / authors)", None),
        ("     hallucinated         — citation does not exist in the PDF", None),
        ("", None),
        ("4. Fill human_is_academic / academic_subfield / is_canonical / citation_function", None),
        ("   ONLY when you disagree with the LLM. Blank = agree.", None),
        ("", None),
        ("5. Scan the PDF for missed citations. Log each in a fn_blank row (yellow):", None),
        ("     verdict = human_found, fill title / year / authors.", None),
        ("", None),
        ("6. Save. Then run: python src/validation/analyze.py", None),
        ("", None),
        ("What counts as a citation", Font(bold=True, size=14)),
        ("Any reference a human reader would recognize: bibliography entries, inline", None),
        ("name-year mentions, informal mentions ('the Fama-French model'), data sources", None),
        ("(CRSP, Compustat), and self-citations. All count.", None),
        ("", None),
        ("Row colors", Font(bold=True, size=14)),
        ("   Light blue   — pre-filled LLM extraction (judge it)", None),
        ("   Light yellow — blank row for logging a missed citation", None),
        ("   Dark gray    — separator between documents", None),
    ]
    for i, (txt, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=txt)
        if font is not None:
            cell.font = font
    ws.column_dimensions["A"].width = 110


def _write_doc_block(ws, start_row, doc_row, llm_citations):
    r = start_row

    for idx, cit in enumerate(llm_citations.to_dict("records"), start=1):
        values = {
            "doc_idx": doc_row["doc_idx"], "doc_id": doc_row["doc_id"],
            "row_type": "llm", "citation_idx": idx,
            "source_title": _safe(doc_row.get("source_title", "")),
            "source_file": doc_row["source_file"],
            "raw_mention":       _safe(cit.get("raw_mention", "")),
            "recovered_title":   _safe(cit.get("recovered_title", "")),
            "recovered_authors": _safe(cit.get("recovered_authors", "")),
            "recovered_year":    _safe(cit.get("recovered_year", "")),
            "recovered_journal": _safe(cit.get("recovered_journal", "")),
            "is_academic":       _safe(cit.get("is_academic", "")),
            "is_canonical":      _safe(cit.get("is_canonical", "")),
            "academic_subfield": _safe(cit.get("academic_subfield", "")),
            "citation_function": _safe(cit.get("citation_function", "")),
            "citation_polarity": _safe(cit.get("citation_polarity", "")),
            "confidence":        _safe(cit.get("confidence", "")),
        }
        for col_idx, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=values.get(h, ""))
            cell.fill = LLM_FILL if col_idx not in INPUT_COLUMNS else INPUT_FILL
        r += 1

    for j in range(FN_BLANKS_PER_DOC):
        values = {
            "doc_idx": doc_row["doc_idx"], "doc_id": doc_row["doc_id"],
            "row_type": "fn_blank", "citation_idx": f"fn_{j+1}",
            "source_title": _safe(doc_row.get("source_title", "")),
            "source_file": doc_row["source_file"],
        }
        for col_idx, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=values.get(h, ""))
            cell.fill = FN_FILL if col_idx not in INPUT_COLUMNS else INPUT_FILL
        r += 1

    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(row=r, column=col_idx, value="").fill = SEPARATOR_FILL
    r += 1
    return r


def _build_audit(ws, sample_df, citations_df):
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.freeze_panes = "G2"

    col_widths = {
        "doc_idx": 8, "doc_id": 14, "row_type": 12, "citation_idx": 12,
        "source_title": 30, "source_file": 40,
        "raw_mention": 40, "recovered_title": 35, "recovered_authors": 25,
        "recovered_year": 10, "recovered_journal": 25,
        "is_academic": 12, "is_canonical": 12, "academic_subfield": 20,
        "citation_function": 18, "citation_polarity": 18, "confidence": 12,
        "human_verdict": 20,
        "human_correct_title": 30, "human_correct_year": 12, "human_correct_authors": 25,
        "human_is_academic": 15, "human_academic_subfield": 20,
        "human_is_canonical": 15, "human_citation_function": 20,
        "notes": 40,
    }
    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(h, 15)

    row = 2
    for _, doc_row in sample_df.iterrows():
        doc_cits = citations_df[citations_df["doc_id"] == doc_row["doc_id"]]
        row = _write_doc_block(ws, row, doc_row, doc_cits)
    last_row = row

    verdict_col   = get_column_letter(HEADERS.index("human_verdict") + 1)
    subfield_col  = get_column_letter(HEADERS.index("human_academic_subfield") + 1)
    function_col  = get_column_letter(HEADERS.index("human_citation_function") + 1)
    academic_col  = get_column_letter(HEADERS.index("human_is_academic") + 1)
    canonical_col = get_column_letter(HEADERS.index("human_is_canonical") + 1)

    dv_verdict  = DataValidation(type="list", formula1=f'"{",".join(VERDICT_OPTIONS)}"', allow_blank=True)
    dv_subfield = DataValidation(type="list", formula1=f'"{",".join(SUBFIELD_OPTIONS)}"', allow_blank=True)
    dv_function = DataValidation(type="list", formula1=f'"{",".join(FUNCTION_OPTIONS)}"', allow_blank=True)
    dv_bool     = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)

    for dv in (dv_verdict, dv_subfield, dv_function, dv_bool):
        ws.add_data_validation(dv)
    dv_verdict.add(f"{verdict_col}2:{verdict_col}{last_row}")
    dv_subfield.add(f"{subfield_col}2:{subfield_col}{last_row}")
    dv_function.add(f"{function_col}2:{function_col}{last_row}")
    dv_bool.add(f"{academic_col}2:{academic_col}{last_row}")
    dv_bool.add(f"{canonical_col}2:{canonical_col}{last_row}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Prepare validation audit: sample + template in one pass")
    parser.add_argument("--n", type=int, default=DEFAULT_N, help=f"Target sample size (default {DEFAULT_N})")
    parser.add_argument("--seed", type=int, default=RANDOM_SEED, help=f"Random seed (default {RANDOM_SEED})")
    args = parser.parse_args()

    if not DOC_PATH.exists():
        raise FileNotFoundError(f"{DOC_PATH} not found. Run src/01_extract_deep.py first.")
    if not CIT_PATH.exists():
        raise FileNotFoundError(f"{CIT_PATH} not found. Run src/01_extract_deep.py first.")

    # --- Sample ---
    docs = pd.read_csv(DOC_PATH, low_memory=False)
    n_auditable = (docs["llm_status"] == "ok").sum()
    print(f"Loaded {len(docs)} documents ({n_auditable} auditable)")
    if n_auditable < args.n:
        print(f"[warn] requested n={args.n} but only {n_auditable} auditable — capping")

    sample = build_sample(docs, args.n, args.seed)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    output_cols = [
        "doc_idx", "doc_id", "source_file", "source_title", "source_type",
        "source_year_final", "source_year", "pdf_extraction_method",
        "doc_text_strategy", "doc_text_truncated", "doc_word_count",
        "doc_page_count", "doc_citation_count", "stratum",
    ]
    cols_present = [c for c in output_cols if c in sample.columns]
    sample[cols_present].to_csv(SAMPLE_PATH, index=False)
    print(f"\nSampled {len(sample)} documents → {SAMPLE_PATH.name}")
    print("  By source_type:", dict(sample["source_type"].value_counts()))
    print("  By extraction :", dict(sample["pdf_extraction_method"].value_counts()))
    print("  Truncated     :", int(sample["doc_text_truncated"].astype(bool).sum()))

    # --- Template ---
    citations = pd.read_csv(CIT_PATH, low_memory=False)
    sampled_cits = citations[citations["doc_id"].isin(sample["doc_id"])].copy()
    print(f"\nLLM citations in sample: {len(sampled_cits)}")
    print(f"FN blank rows per doc:   {FN_BLANKS_PER_DOC}")

    wb = Workbook()
    ws_inst = wb.active
    ws_inst.title = "instructions"
    _build_instructions(ws_inst)
    ws_audit = wb.create_sheet("audit")
    _build_audit(ws_audit, sample, sampled_cits)
    wb.save(XLSX_PATH)
    print(f"\nWrote {XLSX_PATH}")
    print(f"Next: fill it in, then run  python src/validation/analyze.py")


if __name__ == "__main__":
    main()
